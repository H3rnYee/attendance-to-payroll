import openpyxl
import pandas as pd
import re
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import column_index_from_string, get_column_letter




# Load the workbook with calculated values
file_path = "/Users/gohhernyee/Desktop/F10W Probe Tech_Attendance JAN 2025_31 Jan 2025.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

# Select the correct sheet
sheet_name = "DAY "  # Change this to your actual sheet name
if sheet_name not in wb.sheetnames:
    raise ValueError(f"Sheet '{sheet_name}' not found! Available sheets: {wb.sheetnames}")

sheet = wb[sheet_name]

#Take the Individual Rates from Each work permit employee
file_path_WP = "/Users/gohhernyee/Desktop/JAN2025_Payroll_WP_Records_R1_7 Feb 2025.xlsx"
wb1 = openpyxl.load_workbook(file_path_WP, data_only=True)

sheet_name_WP = "JAN"
if sheet_name_WP not in wb1.sheetnames:
    raise ValueError(f"Sheet '{sheet_name_WP}' not found! Available sheets: {wb1.sheetnames}")
    
sheet_rate_WP = wb1[sheet_name_WP]
num_filled_WP = 0
for row in range(2, sheet_rate_WP.max_row + 1):
    cell_value = sheet_rate_WP.cell(row=row, column=2).value
    if cell_value is None or str(cell_value).strip() == "":  # Stop when an empty cell is found
        break
    num_filled_WP += 1

#Take the Individual Rates from Each employee
file_path_EMPLOYEE = "/Users/gohhernyee/Desktop/JAN2025_Payroll_Employees_Records_R1_11 Feb 2025.xlsx"
wb2 = openpyxl.load_workbook(file_path_EMPLOYEE, data_only=True)
sheet_name_EMPLOYEE = "JAN"
if sheet_name_EMPLOYEE not in wb2.sheetnames:
    raise ValueError(f"Sheet '{sheet_name_EMPLOYEE}' not found! Available sheets: {wb2.sheetnames}")

sheet_rate_EMPLOYEE = wb2[sheet_name_EMPLOYEE]
num_filled_employee = 0
for row in range(2, sheet_rate_EMPLOYEE.max_row + 1):
    cell_value = sheet_rate_EMPLOYEE.cell(row=row, column=2).value
    if cell_value is None or str(cell_value).strip() == "":  # Stop when an empty cell is found
        break
    num_filled_employee += 1

#Get last column with name
for col in range(3, sheet.max_column + 1 ):
    # Check if cell is merged
    employee_name = sheet.cell(row=3, column=col).value 
    cell_address = f"{get_column_letter(col)}3" 
    is_merged = any(cell_address in merged_range for merged_range in sheet.merged_cells.ranges)
    if employee_name and not is_merged:
        last_col = col  # Update the last column with a valid employee name



# Payroll output file
payroll_file = "Payroll_Generated.xlsx"
payroll_wb = openpyxl.Workbook()
payroll_wb.remove(payroll_wb.active)  # Remove the default sheet

# Loop through each employee column (starting from column C)
for col in range(3, last_col + 1):  # Assuming data starts from column C
    cell_value = sheet.cell(row=42, column=col).value
    if not cell_value:
        continue  # Skip empty columns
    work_level = sheet.cell(row = 5, column = col).value
    work_level = work_level.strip()

    cell_value = str(cell_value).strip()
    lines = [line.strip() for line in cell_value.split("\n") if line.strip()]

    # Extract values
    # Extract total count and max count from the first line
    # Extract total count (including possible PH) and max count
    total_status = re.search(r"\[(\d+)\*?\/(\d+)\]", lines[0])
    prorated_status = re.search(r"Pro-rated Basic Monthly \[(\d+)\/(\d+)\]", lines[1])
    if prorated_status:
        total_count = int(prorated_status.group(1))
        max_count = int(prorated_status.group(2))
    else:
        total_count = int(total_status.group(1).strip())
        max_count = int(total_status.group(2))

    # Find the line that contains "Non-Shift" dynamically
    non_shift_count = 0  # Default to 0 if not found
    for line in lines:
        match = re.search(r"(\d+)\sNon-Shift", line)
        if match:
            non_shift_count = int(match.group(1))
            break  # Stop searching once found

    # Find the line that contains "Holiday_Shift" dynamically
    holiday_shift_count = 0  # Default to 0 if not found
    for line in lines:
        match = re.search(r"(\d+)\sHoliday_Shift", line)
        if match:
            holiday_shift_count = int(match.group(1))
            break  # Stop searching once found
    # Find the line that contains "Holiday_Non-Shift" dynamically
    holiday_non_shift_count = 0  # Default to 0 if not found
    for line in lines:
        match = re.search(r"(\d+)\sHoliday_Non-Shift", line)
        if match:
            holiday_non_shift_count = int(match.group(1))
            break  # Stop searching once found
    # Find the line that contains "Paid Leave" dynamically
    paid_leave_count = 0  # Default to 0 if not found
    for line in lines:
        match = re.search(r"(\d+)\sPaid Leave", line)
        if match:
            paid_leave_count = int(match.group(1))
            break  # Stop searching once found


    # Compute incentives
    hardshift_incentive = (total_count / 15) * 8 if total_count else 0
    attendance_incentive = 1 if total_count == max_count else 0
    
    # Get employee name from row 3 (above attendance data)
    employee_name = sheet.cell(row=3, column=col).value or f"Employee_{col}"
    # Remove invalid characters from the sheet name
    employee_name = re.sub(r'[\/\\\*\?\[\]\:]', '_', employee_name)  # Replace invalid characters with '_'
    strip_employee_name = employee_name.strip()

    if work_level == "WP EMPLOYEE":
        #Find name of WP, take daily rate from same row 
        for row in range(2, num_filled_WP + 1 ):
            cell_value = sheet_rate_WP.cell(row = row ,column = 2).value
            cell_value_name = re.sub(r'[\/\\\*\?\[\]\:]', '_', cell_value).strip()
            if cell_value_name == strip_employee_name:
                daily_rate = sheet_rate_WP.cell(row = row, column = 6).value
                basic_monthly = sheet_rate_WP.cell(row=row, column = 5).value
                non_shift_amount = non_shift_count * daily_rate * 1.5
                holiday_shift_amount = holiday_shift_count * daily_rate
                holiday_non_shift_amount = holiday_non_shift_count * daily_rate * 2
                save_row = row
                if total_count == max_count:
                    NSK = 100
                else:
                    NSK = 0

                payroll_data = [
                ["", "Unit", "", "Rate", "", "Total"],
                ["Basic Salary", "", "", "", "", basic_monthly],
                ["Non-Shift", non_shift_count, "", daily_rate * 1.5, "per day", non_shift_amount],
                ["Holiday - Shift", holiday_shift_count, "", daily_rate, "per day", holiday_shift_amount],
                ["Holiday - Non-Shift", holiday_non_shift_count, "", daily_rate * 2, "per day", holiday_non_shift_amount],
                ["Hardship Incentive", total_count, "", 8.00, "", total_count * 8],
                ["Attendance Incentive", attendance_incentive, "", 100.00, "", attendance_incentive*100],
                ["NSK Incentive", "", "", "", "", NSK],
                ["IAR Incentive", "", "", "", "", "INSERT IAR INCENTIVE"],
                ["ABR Incentive", "", "", "", "", "INSERT ABR INCENTIVE"],
                ["Gross Total", "", "", "", "", "=SUM(F11:F21)"]
    ]
                break
        print(strip_employee_name,work_level)
    if work_level == "EMPLOYEE":
        for row in range(2, num_filled_employee + 1 ):
            cell_value = sheet_rate_EMPLOYEE.cell(row = row ,column = 2).value
            cell_value_name = re.sub(r'[\/\\\*\?\[\]\:]', '_', cell_value).strip()
            if cell_value_name == strip_employee_name:
                daily_rate = sheet_rate_EMPLOYEE.cell(row = row, column = 8).value
                Shift = daily_rate * total_count
                non_shift_amount = non_shift_count * daily_rate * 1.5
                holiday_shift_amount = holiday_shift_count * daily_rate
                holiday_non_shift_amount = holiday_non_shift_count * daily_rate * 2
                save_row = row
                if sheet_name == "DAY ":
                    shift_time = "Shift Incentive - Day  "
                    shift_incentive = 5
                if sheet_name == "NIGHT":
                    shift_time = "Shift Incentive - Night  "
                    shift_incentive = 11
                shift_incentive_pay = shift_incentive * total_count
                payroll_data = [
                ["", "Unit", "", "Rate", "", "Total"],
                ["Shift", total_count, "", daily_rate, "", Shift],
                ["Non-Shift", non_shift_count, "", daily_rate * 1.5, "per day", non_shift_amount],
                ["Holiday - Shift", holiday_shift_count, "", daily_rate, "per day", holiday_shift_amount],
                ["Holiday - Non-Shift", holiday_non_shift_count, "", daily_rate * 2, "per day", holiday_non_shift_amount],
                ["Hardship Incentive", total_count, "", 8.00, "", total_count * 8],
                ["Attendance Incentive", attendance_incentive, "", 100.00, "", attendance_incentive*100],
                ["Paid Leave",paid_leave_count,"", daily_rate, "per day" , paid_leave_count * daily_rate],
                [shift_time, total_count, "", shift_incentive, "", shift_incentive_pay],
                ["Gross Total", "", "", "", "", "=SUM(F11:F18)"]
    ]
                break
        print(strip_employee_name,work_level)


    
    # Create a new sheet for the employee
    ws = payroll_wb.create_sheet(title=employee_name)
    
    # Insert payroll data
    for row_idx, row in enumerate(payroll_data, start=10):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    if work_level == "WP EMPLOYEE":
        # Bold the Information
        for row in ws.iter_rows(min_row=2, max_row=7, min_col=1, max_col=5):
            for cell in row:
                cell.font = Font(bold=True, name ="Calibri", size = 11) 
        # Format headers
        ws.sheet_view.showGridLines = False
        ws.merge_cells("A1:G1")
        ws["A1"] = "PAY SLIP"
        ws["A1"].font = Font(name="Calibri", bold=True, size=22)
        ws["A1"].alignment = Alignment(horizontal="center")
        NRIC = sheet_rate_WP.cell(row = save_row, column = 4).value
        # Employee details
        ws["A2"] = f"Name of Employee : {employee_name}"

        ws["A3"] = f"NRIC No. : {NRIC}"

        ws["A4"] = f"FIN : {NRIC}"

        ws["A5"] = f"Work Permit No. : {NRIC}"

        bank = sheet_rate_WP.cell(row = save_row, column = 12).value
        ws["A6"] = f"Bank : {bank}"

        bank_account = sheet_rate_WP.cell(row = save_row, column = 13).value
        ws["A7"] = f"Bank Accoutn No. : {bank_account}"

        ws["E2"] = "Designation :  "
        ws["F2"] = "=JAN!C20"
        ws["E3"] = "Service Period :"
        ws["F3"] = "=JAN!C51"
        ws["E4"] = "Date of Payment :"
        ws["F4"] = "=JAN!C52"
        ws["E5"] = "Mode of Payment : "
        ws["F5"] = "Bank Transfer"

    else:
        NRIC = sheet_rate_EMPLOYEE.cell(row = save_row, column = 4).value

    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

# Save the final payroll file
payroll_wb.save(payroll_file)
print(f"Payroll generated successfully: {payroll_file}")
